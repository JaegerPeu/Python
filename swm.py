# Dashboard SWM MFO
import streamlit.components.v1 as components 
import streamlit as st
from openpyxl import load_workbook
import pandas as pd

# === CONFIGURAÇÕES ===
st.set_page_config(page_title="Dashboard Back Office", layout="wide")
st.title("Controle SWM | MFO")

# === CAMINHO E ABA ===
caminho_arquivo = "Números MFO 31072025.xlsx"
aba = "Total"

try:
    wb = load_workbook(caminho_arquivo, data_only=True)
    if aba not in wb.sheetnames:
        st.error(f"Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
        st.stop()
    ws = wb[aba]
except Exception as e:
    st.error(f"Erro ao carregar planilha: {e}")
    st.stop()

# === FORMATADORES ===
def formatar(valor):
    if valor is None or pd.isna(valor):
        return "N/A"
    try:
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return str(valor)

def formatar_off(valor):
    if valor is None or pd.isna(valor):
        return "N/A"
    try:
        return f"US$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return str(valor)

def formatar_reais_str(valor):
    if valor is None or pd.isna(valor):
        return "N/A"
    try:
        if isinstance(valor, str):
            valor = valor.replace("R$", "").replace(".", "").replace(",", ".")
        return formatar(float(valor))
    except:
        return str(valor)

# === LER DADOS ===
pl_onshore          = ws["F5"].value
pl_offshore         = ws["F8"].value
pl_offshore_brl     = ws["F10"].value
pl_total            = ws["F13"].value
receita_onshore     = ws["B5"].value
receita_offshore    = ws["B8"].value
receita_total       = ws["B13"].value
imposto             = receita_total * 0.15
custos              = ws["B17"].value * 12
lucro               = ws["B18"].value * 12
fidc                = ws["B20"].value

# === ABAS ===
aba1, aba2 = st.tabs(["📊 Visão Geral", "📂 Veículos de Alocação"])

# === ABA 1: Visão Geral ===
with aba1:
    col1, col2, col3 = st.columns(3)
    col1.metric("PL Onshore", formatar(pl_onshore))
    col2.metric("PL Offshore", formatar_off(pl_offshore))
    col3.metric("PL Total", formatar(pl_total))

    col4, col5, col6 = st.columns(3)
    col4.metric("Receita Est. Onshore", formatar(receita_onshore))
    col5.metric("Receita Est. Offshore", formatar_off(receita_offshore))
    col6.metric("Receita Est. Bruta Total Ano", formatar(receita_total))

    st.divider()
    st.subheader("Resultado Final")
    col7, col8, col9 = st.columns(3)
    col7.metric("Imposto", formatar(imposto))
    col8.metric("Custo Fixo", formatar(custos))
    col9.metric("Lucro", formatar(lucro))
    
    st.divider()
    st.subheader("Outros Dados")
    st.metric("FIDC TAMBASA", formatar(fidc))

# === ABA 2: Veículos de Alocação ===
with aba2:
    st.subheader("📁 Detalhamento dos Fundos e Veículos")

    df = pd.read_excel(caminho_arquivo, sheet_name="Veiculos")

    html_style = """
    <style>
    .card-container {
        display: flex;
        flex-wrap: wrap;
        justify-content: center;
    }
    .card {
        border-radius: 10px;
        padding: 15px;
        margin: 10px;
        font-family: Arial, sans-serif;
        width: 310px;
        box-shadow: 0px 2px 8px rgba(0,0,0,0.1);
    }
    .header {
        font-weight: bold;
        font-size: 16px;
        margin-bottom: 5px;
        text-align: center;
    }
    .meta {
        font-size: 13px;
        color: #333;
        margin-bottom: 4px;
        text-align: center;
    }
    </style>
    """

    html_cards = '<div class="card-container">'
    for _, row in df.iterrows():
        fundo = row['Fundo']
        grande_classe = row['Grande Classe']
        cor_fundo = "#e8f5e9" if grande_classe == "FI" else "#e3f2fd"

        try:
            aum_percent = f"{float(row['%AUM']) * 100:.2f}%"
        except:
            aum_percent = str(row['%AUM'])

        try:
            taxa_adm = f"{float(row['Taxa Adm']) * 100:.2f}%"
        except:
            taxa_adm = str(row['Taxa Adm'])

        try:
            te_percent = f"{float(row['Te']) * 100:.2f}%"
        except:
            te_percent = str(row['Te'])

        patrimonio = formatar_reais_str(row['Patrimônio'])
        receita_total = formatar_reais_str(row['Receita Total Estimada'])
        receita_asset = formatar_reais_str(row['Receita Asset Estimada'])

        html_cards += f"""
        <div class="card" style="background-color:{cor_fundo}; border-left: 5px solid #666;">
            <div class="header">📌 {fundo}</div>
            <div class="meta">🏦 Classe: {grande_classe} | 💼 {row['Custodiante']}</div>

            <br>
            <div style="display: flex; justify-content: space-between;">
                <div style="text-align:center;">
                    <div style="font-size: 20px; font-weight: bold;">{patrimonio}</div>
                    <div style="font-size: 12px; color: #666;">Patrimônio</div>
                </div>
                <div style="text-align:center;">
                    <div style="font-size: 20px; font-weight: bold;">{aum_percent}</div>
                    <div style="font-size: 12px; color: #666;">% AUM</div>
                </div>
                <div style="text-align:center;">
                    <div style="font-size: 20px; font-weight: bold;">{taxa_adm}</div>
                    <div style="font-size: 12px; color: #666;">Taxa Adm</div>
                </div>
            </div>
            <br>
            <hr style="margin: 10px 0;">
            <div class="meta">👤 Gestor: {row['Gestor']}</div>
            <div class="meta">💸 Receita Total Estimada: {receita_total}</div>
            <div class="meta">🏦 Receita Asset Estimada: {receita_asset}</div>
        </div>
        """
    html_cards += '</div>'

    altura_estimativa = max(600, len(df) * 420)
    components.html(html_style + html_cards, height=altura_estimativa, scrolling=False)

